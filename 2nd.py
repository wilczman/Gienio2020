import openpyxl


def choose():
    while True:
        inp = input(f'Jaki produkt? ')
        if inp.isnumeric():
            print('TODO')
            continue
        elif inp.isalpha():
            if 'ex' in inp:
                exit(0)
            else:
                return inp
        else:
            continue


def type_amount():
    while True:
        inp = input(f'Ile (kg) potrzebujesz? ')
        if inp.isnumeric() and 0 < int(inp) :
            return int(inp)
        elif 'ex' in inp:
            exit(0)
        else:
            continue


def search(inp):
    wb = openpyxl.load_workbook('Przepis.xlsx')
    tab_of_sheets = wb.sheetnames
    for name in tab_of_sheets:
        if inp in name.lower():
            print("Wybrano produkt: " + name)
            return name
        else:
            continue
    print('Nie znaleziono pozycji!')


while True:
    name = 0
    while name == 0:
        choice = choose()
        name = search(choice)
    amount = type_amount()
    wb = openpyxl.load_workbook('Przepis.xlsx')
    sheet = wb[name]
    print('Potrzebujesz: ')
    for row in range(2, sheet.max_row):
        print(f"{sheet['A'+ str(row)].value}: {round(sheet['B'+ str(row)].value/100 * amount, 3)}{sheet['C'+ str(row)].value}")

