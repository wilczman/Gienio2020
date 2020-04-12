import tkinter as tk
import fileWithChoosing
import openpyxl
import testPrint


RECIPE_FILE_NAME = 'Przepis.xlsx'
HEIGHT = 700
WIDTH = 800
AMOUNT = 200    # TODO


def get_position(string):
    # print(f"This is entry: {string}")
    wb = openpyxl.load_workbook(RECIPE_FILE_NAME)
    if string.isnumeric():  # TODO
        format_response('numeric is currently unavailable')
    elif string == '':
        format_response('Nie wpisałeś pozycji!')
        entry_name.delete(0, 'end')
        entry_amount.place_forget()
        print_button.place_forget()
    elif fileWithChoosing.search(string, wb) == 0:
        format_response("Nie znaleziono pozycji!")
        entry_name.delete(0, 'end')
        entry_amount.place_forget()
        print_button.place_forget()
    else:
        name = fileWithChoosing.search(string, wb)
        sheet = wb[name]
        format_response(f'Wybrano produkt: {name}\n\nWprowadź ilość')
        amount_button.place(relx=0.7, rely=0, relwidth=0.3, relheight=1)
        entry_amount.place(relwidth=0.6, relheight=1)


def get_amount(number, name):
    if number.isnumeric():
        try:
            wb = openpyxl.load_workbook(RECIPE_FILE_NAME)
            name = fileWithChoosing.search(name, wb)
            sheet = wb[name]
            output = ''
            output += f'Na {number} kg produktu: {name}, potrzebujesz: \n\n'
            for row in range(2, sheet.max_row+1):
                output += f"{sheet['A' + str(row)].value}:      {round(sheet['B' + str(row)].value / 100 * int(number), 2)} {sheet['C' + str(row)].value}\n"
                format_response(output)
            amount_button.place_forget()
            entry_amount.place_forget()
            entry_name.delete(0, 'end')
            entry_amount.delete(0, 'end')
            print_button.place(relx=0.3, rely=0, relwidth=0.4, relheight=0.9)
        except:
            format_response('Błąd odczytu, wpisz od nowa nazwę!')
            amount_button.place_forget()
            entry_amount.place_forget()
            print_button.place_forget()
            entry_name.delete(0, 'end')
    else:
        format_response('Zła ilość!')
        amount_button.place_forget()
        entry_amount.place_forget()
        print_button.place_forget()
        entry_name.delete(0, 'end')
        entry_amount.delete(0, 'end')


def printer():
    testPrint.printer(label['text'])


def format_response(text='default'):
    label['text'] = text


if __name__ == "__main__":
    root = tk.Tk()
    root.title('Gienio2020')
    canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)
    canvas.pack()

    name_frame = tk.Frame(root, bg='#80c1ff', bd=5)
    name_frame.place(relwidth=1, relheight=0.1, relx=0.5, rely=0, anchor='n')

    amount_frame = tk.Frame(root, bg='#80c1ff', bd=5)
    amount_frame.place(relwidth=1, relheight=0.1, relx=0.5, rely=0.1, anchor='n')

    lower_frame = tk.Frame(root, bg='#80c1ff', bd=10)
    lower_frame.place(relwidth=1, relheight=0.7, relx=0.5, rely=0.2, anchor='n')

    print_frame = tk.Frame(root, bg='#80c1ff', bd=1)
    print_frame.place(relwidth=1, relheight=0.1, relx=0.5, rely=0.9, anchor='n')

    entry_name = tk.Entry(name_frame, bg='white', font=40)
    # entry.pack(side='left', fill='both')
    entry_name.place(relwidth=0.6, relheight=1)

    entry_amount = tk.Entry(amount_frame, bg='white', font=40)

    amount_button = tk.Button(amount_frame, text="Oblicz", bg='white', fg='black',
                       font=20, command=lambda: get_amount(entry_amount.get(), entry_name.get()))

    name_button = tk.Button(name_frame, text="Szukaj", bg='white', fg='black',
                       font=20, command = lambda: get_position(entry_name.get()))
    name_button.place(relx=0.7, rely=0, relwidth=0.3, relheight=1)

    print_button = tk.Button(print_frame, text="Drukuj", bg='white', fg='black',
                              font=20, command=lambda: printer())


    # button.pack(side='left', fill='both', expand='True')
    # button.grid(row=0, column=0)

    # label = tk.Label(frame, text="This is a label", bg='white')
    # label.pack(side='left', fill='both')
    # label.place(relx=0.3, rely=0, relwidth=0.45, relheight=0.25)

    label = tk.Label(lower_frame, bg='white', font=15)
    label.place(relwidth=1, relheight=1)

    root.mainloop()




